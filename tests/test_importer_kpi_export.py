import tempfile
import unittest
from datetime import date
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cmit.config import PROJECT_ROOT, TEMPLATE_PATH
from cmit.delay_codes import is_deductible_code
from cmit.importer import normalize_move_kind, read_move_file
from cmit.kpi import calculate_report
from cmit.report_exporter import export_shift_report
from cmit.shift import make_shift
from cmit.storage import default_manual_inputs


class ImportKpiExportTest(unittest.TestCase):
    def test_move_kind_normalization(self):
        self.assertEqual(normalize_move_kind("Receival"), "GATE IN")
        self.assertEqual(normalize_move_kind("Delivery"), "GATE OUT")
        self.assertEqual(normalize_move_kind("Yard Move"), "HOUSE KEEPING")
        self.assertEqual(normalize_move_kind("Discharge"), "DISCHARGE")
        self.assertTrue(is_deductible_code("CODE 52"))
        self.assertFalse(is_deductible_code("CODE 31"))

    def test_only_code_5x_reduces_gmph(self):
        shift = make_shift(date(2026, 5, 26), "D1")
        moves = pd.DataFrame(
            [
                {
                    "UNIT NBR": "A",
                    "MOVE KIND": "LOAD",
                    "CARRIER VISIT": "TESTV",
                    "LOAI PTVT": "VESSEL",
                    "LOAI CAU BEN": "CMIT QUAY",
                    "STS QUAY": "QC01",
                    "TIME STS QUAY": datetime(2026, 5, 26, 6, 0),
                },
                {
                    "UNIT NBR": "B",
                    "MOVE KIND": "LOAD",
                    "CARRIER VISIT": "TESTV",
                    "LOAI PTVT": "VESSEL",
                    "LOAI CAU BEN": "CMIT QUAY",
                    "STS QUAY": "QC01",
                    "TIME STS QUAY": datetime(2026, 5, 26, 7, 0),
                },
            ]
        )
        manual = default_manual_inputs()
        manual["delays"] = [{"carrier_visit": "TESTV", "crane": "QC01", "code": "CODE 31", "minutes": 30}]
        non_deduct = calculate_report(moves, manual, shift).vessel_rows[1]
        self.assertEqual(non_deduct["net_hours"], 1.0)
        manual["delays"] = [{"carrier_visit": "TESTV", "crane": "QC01", "code": "CODE 52", "minutes": 30}]
        deduct = calculate_report(moves, manual, shift).vessel_rows[1]
        self.assertEqual(deduct["net_hours"], 0.5)

    def test_import_raw_n4_and_calculate(self):
        sample = PROJECT_ROOT / "MoveEvent_20260526_2203.xlsx"
        if not sample.exists():
            self.skipTest("Sample MoveEvent file is not available.")
        shift = make_shift(date(2026, 5, 26), "D1")
        moves = read_move_file(sample, shift)
        self.assertGreater(len(moves), 0)
        self.assertIn("MOVE KIND", moves.columns)
        bundle = calculate_report(moves, default_manual_inputs(), shift)
        self.assertGreaterEqual(bundle.summary_rows[0]["overall_volume"], 0)

    def test_export_workbook_has_required_sheets(self):
        sample = PROJECT_ROOT / "MoveEvent_20260526_2203.xlsx"
        if not sample.exists():
            self.skipTest("Sample MoveEvent file is not available.")
        shift = make_shift(date(2026, 5, 26), "D1")
        moves = read_move_file(sample, shift)
        bundle = calculate_report(moves, default_manual_inputs(), shift)
        with tempfile.TemporaryDirectory() as tmp:
            output = export_shift_report(bundle, template_path=TEMPLATE_PATH, output_dir=Path(tmp))
            self.assertTrue(output.exists())
            wb = load_workbook(output, data_only=True)
            self.assertEqual(
                wb.sheetnames,
                ["SUMMARY REPORT", "CRANE STATS", "VESSEL REPORT", "BARGE REPORT", "VESSEL NOTES", "EQUIPMENT BREAKDOWN", "DATA"],
            )
            self.assertEqual(wb["SUMMARY REPORT"]["A1"].value, "SHIFT OPERATION REPORT")


if __name__ == "__main__":
    unittest.main()
