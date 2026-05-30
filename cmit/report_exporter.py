from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import OUTPUT_DIR, TEMPLATE_PATH, ensure_runtime_dirs
from .importer import DATA_COLUMNS
from .kpi import CHE_TYPES, REPORT_LOCATIONS, ReportBundle


EXPORT_DATA_COLUMNS = [c for c in DATA_COLUMNS if c != "TEU"]


def export_shift_report(bundle: ReportBundle, template_path: Path | None = None, output_dir: Path | None = None) -> Path:
    ensure_runtime_dirs()
    template = template_path or TEMPLATE_PATH
    output_dir = output_dir or OUTPUT_DIR
    if template.exists():
        wb = load_workbook(template)
    else:
        wb = _create_fallback_workbook()

    _write_summary(wb["SUMMARY REPORT"], bundle)
    _write_crane_stats(wb["CRANE STATS"], bundle)
    _write_detail_sheet(wb["VESSEL REPORT"], bundle.vessel_rows)
    _write_detail_sheet(wb["BARGE REPORT"], bundle.barge_rows)
    _write_vessel_notes(wb["VESSEL NOTES"], bundle)
    _write_equipment_breakdown(wb["EQUIPMENT BREAKDOWN"], bundle)
    _write_data_sheet(wb["DATA"], bundle)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"SHIFT_REPORT_{bundle.shift.file_label}.xlsx"
    wb.save(output_path)
    return output_path


def _write_summary(ws, bundle: ReportBundle) -> None:
    shift = bundle.shift
    manual = bundle.manual
    personnel = manual.get("personnel", {})
    ws["A1"] = "SHIFT OPERATION REPORT"
    ws["A2"] = f"DATE: {shift.work_date.isoformat()} / SHIFT: {shift.code}"
    ws["B6"] = personnel.get("shift_manager", "")
    ws["B7"] = personnel.get("vessel_supervisors", "")
    ws["B8"] = personnel.get("yard_supervisors", "")

    for idx, row_data in enumerate(bundle.summary_rows, start=12):
        ws.cell(idx, 1, row_data["location"])
        values = [
            row_data["cranes"],
            row_data["vessel_disch"],
            row_data["vessel_load"],
            row_data["vessel_restow"],
            row_data["vessel_total"],
            row_data["vessel_gmph"],
            row_data["barge_disch"],
            row_data["barge_load"],
            row_data["barge_total"],
            row_data["barge_gmph"],
            row_data["overall_volume"],
            row_data["overall_gmph"],
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(idx, col, value)

    cmit = bundle.summary_rows[0] if bundle.summary_rows else {}
    total_values = [
        cmit.get("cranes", "-"),
        cmit.get("vessel_disch", 0),
        cmit.get("vessel_load", 0),
        cmit.get("vessel_restow", 0),
        cmit.get("vessel_total", 0),
        "",
        cmit.get("barge_disch", 0),
        cmit.get("barge_load", 0),
        cmit.get("barge_total", 0),
        "",
        cmit.get("overall_volume", 0),
        "",
    ]
    for col, value in enumerate(total_values, start=2):
        ws.cell(18, col, value)

    gate = bundle.gate
    ws["B21"] = gate.get("gate_in", 0)
    ws["C21"] = gate.get("gate_out", 0)
    ws["D21"] = gate.get("total", 0)
    ws["E21"] = gate.get("ttt", "")
    ws["F21"] = manual.get("other", {}).get("hk_moves") or _auto_hk_moves(bundle)
    ws["G21"] = manual.get("other", {}).get("vessel_sequence") or _auto_vessel_sequence(bundle)
    ws["L21"] = manual.get("other", {}).get("boe", "--")
    ws["M21"] = manual.get("other", {}).get("incident", "NONE")

    che = bundle.che
    for idx, kind in enumerate(CHE_TYPES, start=2):
        ws.cell(24, idx, _display_number(che["demand"].get(kind, 0)))
        ws.cell(25, idx, _display_dash_number(che["bd_time"].get(kind, 0)))
        ws.cell(26, idx, _display_dash_number(che["actual"].get(kind, 0), decimals=2))
        ws.cell(27, idx, f"{che['availability'].get(kind, 1) * 100:.2f}%")
    ws["H24"] = che.get("notes", "")


def _write_crane_stats(ws, bundle: ReportBundle) -> None:
    cmit_stats = bundle.crane_stats.get("CMIT QUAY (CẦU BẾN CMIT)", {})
    metrics = [
        cmit_stats.get("total_moves", 0),
        cmit_stats.get("gross_hours", 0),
        cmit_stats.get("delay_hours", 0),
        cmit_stats.get("net_hours", 0),
        f"{cmit_stats.get('shift_all', 0):.2f}",
        f"{cmit_stats.get('shift_vessel', 0):.2f}",
        f"{cmit_stats.get('shift_barge', 0):.2f}",
    ]
    for row, value in enumerate(metrics, start=2):
        ws.cell(row, 2, value)
        for col in range(3, 8):
            ws.cell(row, col, 0 if row <= 5 else "0.00")


def _write_detail_sheet(ws, rows: list[dict[str, Any]]) -> None:
    header_styles = [_style_snapshot(ws.cell(1, col)) for col in range(1, 15)]
    all_styles = [_style_snapshot(ws.cell(2, col)) for col in range(1, 15)] if ws.max_row >= 2 else header_styles
    detail_styles = [_style_snapshot(ws.cell(3, col)) for col in range(1, 15)] if ws.max_row >= 3 else header_styles
    _clear_body(ws, 1)

    row_idx = 2
    if not rows:
        _write_row(ws, row_idx, ["NO DATA", "-", "", "", 0, 0, 0, 0, 0, 0, 0, "--", "--", ""], all_styles)
        return

    current_group_start = None
    current_carrier = None
    for row in rows:
        is_all = bool(row.get("is_all"))
        if is_all:
            if current_group_start and row_idx - current_group_start > 1:
                ws.merge_cells(start_row=current_group_start, start_column=1, end_row=row_idx - 1, end_column=1)
            current_group_start = row_idx
            current_carrier = row.get("carrier_visit")
        values = [
            row.get("carrier_name", "") if is_all else "",
            row.get("crane", ""),
            row.get("first_lift"),
            row.get("last_lift"),
            row.get("gross_hours", 0),
            row.get("net_hours", 0),
            row.get("moves", 0),
            row.get("disch", 0),
            row.get("load", 0),
            row.get("restow", 0),
            row.get("gmph", 0),
            row.get("pmph", "--"),
            row.get("bmph", "--"),
            row.get("remarks", ""),
        ]
        _write_row(ws, row_idx, values, all_styles if is_all else detail_styles)
        row_idx += 1
    if current_group_start and row_idx - current_group_start > 1:
        ws.merge_cells(start_row=current_group_start, start_column=1, end_row=row_idx - 1, end_column=1)
    _apply_table_border(ws, 1, row_idx - 1, 14)


def _write_vessel_notes(ws, bundle: ReportBundle) -> None:
    _clear_body(ws, 0)
    row = 1
    if not bundle.vessel_notes:
        ws["A1"] = "NO VESSEL NOTES"
        return
    for note in bundle.vessel_notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row, 1, f"{note['header']} - OPERATION NOTES:")
        _header_fill(ws.cell(row, 1))
        fields = [
            ("VESSEL NAME", note.get("vessel_name", "")),
            ("SERVICE", note.get("service", "")),
            ("FIRST LINE", note.get("first_line", "")),
            ("ALL LINE FAST", note.get("all_line_fast", "")),
            ("GANGWAY SECURE", note.get("gangway_secure", "")),
            ("ETC", note.get("etc", "")),
            ("ETD", note.get("etd", "")),
            ("FIRST LIFT", _fmt_dt(note.get("first_lift"))),
            ("LAST LIFT", _fmt_dt(note.get("last_lift"))),
            ("LASHING FINISH", note.get("lashing_finish", "")),
            ("LAST LINE", note.get("last_line", "")),
            ("TOTAL VOLUME [DIS/LOAD/RESTOW]", note.get("total_volume", "")),
            ("AIT (MINS)", _ait_text(note)),
            ("DIT (MINS)", note.get("dit", "")),
            ("GMPH", note.get("gmph", "")),
            ("BMPH", note.get("bmph", "")),
            ("PMPH", note.get("pmph", "")),
            ("MOVES COMPLETED IN SHIFT", note.get("moves_completed", "")),
            ("REMAINING MOVES", note.get("remaining_moves", "")),
            ("SHIFT PRODUCTIVITY", note.get("shift_productivity", "")),
            ("OPERATION NOTES", note.get("operation_notes", "")),
        ]
        for label, value in fields:
            row += 1
            ws.cell(row, 1, label)
            ws.cell(row, 2, value)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
        ws.cell(row, 1, "CHALLENGES")
        ws.cell(row, 2, "RESOLUTION/ SUGGESTIONS")
        ws.cell(row, 3, "FURTHER REMARKS")
        for col in range(1, 4):
            _header_fill(ws.cell(row, col))
        row += 1
        ws.cell(row, 1, note.get("challenges", "NIL"))
        ws.cell(row, 2, note.get("resolution", "NIL"))
        ws.cell(row, 3, note.get("further_remarks", ""))
        row += 3
    _apply_table_border(ws, 1, row - 1, 3)


def _write_equipment_breakdown(ws, bundle: ReportBundle) -> None:
    header_styles = [_style_snapshot(ws.cell(1, col)) for col in range(1, 8)]
    body_styles = [_style_snapshot(ws.cell(2, col)) for col in range(1, 8)] if ws.max_row >= 2 else header_styles
    _clear_body(ws, 1)
    rows = bundle.equipment_breakdowns or []
    if not rows:
        _write_row(ws, 2, ["-", "-", "-", "-", "-", 0, ""], body_styles)
        return
    for row_idx, item in enumerate(rows, start=2):
        _write_row(
            ws,
            row_idx,
            [
                item.get("che", ""),
                item.get("type", ""),
                item.get("fault_description", ""),
                item.get("start", ""),
                item.get("end", ""),
                item.get("duration_min", ""),
                item.get("remarks", ""),
            ],
            body_styles,
        )
    _apply_table_border(ws, 1, len(rows) + 1, 7)


def _write_data_sheet(ws, bundle: ReportBundle) -> None:
    header_styles = [_style_snapshot(ws.cell(1, col)) for col in range(1, len(EXPORT_DATA_COLUMNS) + 1)]
    body_styles = [_style_snapshot(ws.cell(2, col)) for col in range(1, len(EXPORT_DATA_COLUMNS) + 1)] if ws.max_row >= 2 else header_styles
    _clear_body(ws, 0)
    for col, header in enumerate(EXPORT_DATA_COLUMNS, start=1):
        ws.cell(1, col, header)
        _apply_style(ws.cell(1, col), header_styles[col - 1] if col - 1 < len(header_styles) else None)
    df = bundle.moves.copy()
    for col in EXPORT_DATA_COLUMNS:
        if col not in df.columns:
            df[col] = None
    for row_idx, (_, record) in enumerate(df[EXPORT_DATA_COLUMNS].iterrows(), start=2):
        values = [_excel_value(record[col]) for col in EXPORT_DATA_COLUMNS]
        _write_row(ws, row_idx, values, body_styles)


def _clear_body(ws, keep_rows: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row > keep_rows:
            ws.unmerge_cells(str(merged))
    if ws.max_row > keep_rows:
        ws.delete_rows(keep_rows + 1, ws.max_row - keep_rows)


def _write_row(ws, row_idx: int, values: list[Any], styles: list[dict[str, Any]] | None = None) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row_idx, col_idx, _excel_value(value))
        if styles and col_idx <= len(styles):
            _apply_style(cell, styles[col_idx - 1])
        if isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"


def _style_snapshot(cell) -> dict[str, Any]:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
    }


def _apply_style(cell, style: dict[str, Any] | None) -> None:
    if not style:
        return
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]


def _apply_table_border(ws, min_row: int, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color="000000")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "center", vertical="center", wrap_text=True)


def _header_fill(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="BDD7EE")
    cell.font = Font(bold=True)


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _fmt_dt(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m %H:%M")
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime().strftime("%d/%m %H:%M")
    return str(value or "")


def _ait_text(note: dict[str, Any]) -> str:
    ait = str(note.get("ait", "")).strip()
    ait2 = str(note.get("ait2", "")).strip()
    if ait and ait2:
        return f"AIT: {ait} / AIT2: {ait2}"
    if ait:
        return f"AIT: {ait}"
    return ""


def _display_number(value: Any) -> Any:
    try:
        return int(value) if float(value).is_integer() else round(float(value), 2)
    except Exception:
        return value


def _display_dash_number(value: Any, decimals: int = 0) -> Any:
    try:
        numeric = float(value)
        if numeric == 0:
            return "-"
        return f"{numeric:.{decimals}f}" if decimals else int(numeric)
    except Exception:
        return value or "-"


def _auto_hk_moves(bundle: ReportBundle) -> int:
    return int(bundle.moves["MOVE KIND"].isin(["HOUSE KEEPING", "SHIFTING"]).sum()) if not bundle.moves.empty else 0


def _auto_vessel_sequence(bundle: ReportBundle) -> str:
    lines = []
    for idx, note in enumerate(bundle.vessel_notes, start=1):
        ait = note.get("ait", "")
        ait2 = note.get("ait2", "")
        suffix = f" (AIT: {ait}' - AIT2: {ait2}')" if ait or ait2 else ""
        lines.append(f"{idx}. {note.get('header', note.get('carrier_visit', ''))}{suffix}")
    return "\n".join(lines)


def _create_fallback_workbook() -> Workbook:
    wb = Workbook()
    wb.active.title = "SUMMARY REPORT"
    for name in ["CRANE STATS", "VESSEL REPORT", "BARGE REPORT", "VESSEL NOTES", "EQUIPMENT BREAKDOWN", "DATA"]:
        wb.create_sheet(name)
    for ws in wb.worksheets:
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 14
    return wb

